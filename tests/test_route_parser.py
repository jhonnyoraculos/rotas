from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from services.excel_importer import analyze_workbook, build_import_snapshot
from utils.city_normalizer import (
    Municipality,
    normalize_text,
    resolve_municipality_fields,
)
from utils.route_parser import extract_route_code


def sample_workbook() -> BytesIO:
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "CARREGAMENTOS ATUALIZADOS"
    schedule.append(["Relatório de carregamentos"])
    schedule.append(
        ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]
    )
    schedule.append(
        [
            "DIVINÓPOLIS (R.10)",
            "ITAÚNA ( R. 40 )",
            "PARÁ DE MINAS (R.41)",
            "LAVRAS (R.60)",
            "PONTE NOVA (R.401)",
        ]
    )

    cities = workbook.create_sheet("CIDADES X ROTAS ATUALIZADAS")
    cities["A1"] = "ITAÚNA ( R. 40 )"
    cities["A2"] = "ITAÚNA"
    cities["A3"] = "MATEUS LEME"
    cities["A4"] = "JUATUBA"
    cities["A5"] = "PARÁ DE MINAS (R.41)"
    cities["A6"] = "PARA DE MINAS"
    cities["B1"] = "DIVINÓPOLIS (R.10)"
    cities["B2"] = "DIVINÓPOLIS"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def weekday_city_workbook() -> BytesIO:
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "CARREGAMENTOS ATUALIZADOS"
    schedule.append(
        ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]
    )
    schedule.append(["ITAÚNA (R.40)"] * 5)

    cities = workbook.create_sheet("CIDADES X ROTAS ATUALIZADAS")
    cities.append(
        ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]
    )
    cities.append(["ITAÚNA (R.40)"] * 5)
    cities.append(["MATEUS LEME", "AZURITA", "JUATUBA", "ITAÚNA", "MATEUS LEME"])
    cities.append(["JUATUBA", "MATEUS LEME", "AZURITA", None, "AZURITA"])
    cities.append(["PARÁ DE MINAS (R.41)"] * 5)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def test_parser_understands_schedule_and_route_city_blocks() -> None:
    analysis = analyze_workbook(sample_workbook())

    assert analysis.header_row == 2
    assert analysis.schedule[1] == ["R.40"]
    assert analysis.routes["R.40"].cities == ["ITAÚNA", "MATEUS LEME", "JUATUBA"]
    assert analysis.routes["R.41"].cities == ["PARA DE MINAS"]
    assert "R.401" in analysis.routes


def test_parser_preserves_different_cities_for_each_weekday() -> None:
    analysis = analyze_workbook(weekday_city_workbook())

    assert analysis.weekday_routes[0]["R.40"].cities == [
        "MATEUS LEME",
        "JUATUBA",
    ]
    assert analysis.weekday_routes[1]["R.40"].cities == [
        "AZURITA",
        "MATEUS LEME",
    ]
    assert analysis.weekday_routes[3]["R.40"].cities == ["ITAÚNA"]


def test_route_regex_and_accent_normalization() -> None:
    assert extract_route_code("Itaúna ( R. 040 )") == "R.40"
    assert normalize_text("PARÁ   DE MINAS") == normalize_text("Para de Minas")


def test_missing_ibge_code_is_filled_from_official_municipality() -> None:
    municipalities = (
        Municipality("São Sebastião do Oeste", "MG", "3164605"),
    )

    result = resolve_municipality_fields(
        "SÃO SEBASTIÃO DO OESTE",
        "sao sebastiao do oeste",
        "mg",
        "",
        municipalities,
    )

    assert result == ("São Sebastião do Oeste", "MG", "3164605")


def test_ibge_code_is_not_guessed_for_a_different_name() -> None:
    municipalities = (
        Municipality("São Sebastião do Oeste", "MG", "3164605"),
    )

    result = resolve_municipality_fields(
        "SÃO SEBASTIÃO OESTE",
        "São Sebastião Oeste",
        "MG",
        "",
        municipalities,
    )

    assert result == ("São Sebastião Oeste", "MG", "")


def test_missing_ibge_code_falls_back_to_original_city() -> None:
    municipalities = (
        Municipality("Carmo da Cachoeira", "MG", "3113908"),
    )

    result = resolve_municipality_fields(
        "CARMO DA CACHOEIRA",
        "Carmo do Cachoeira",
        "MG",
        "",
        municipalities,
    )

    assert result == ("Carmo da Cachoeira", "MG", "3113908")


def test_route_name_is_included_when_it_is_an_official_municipality(
    monkeypatch,
) -> None:
    analysis = analyze_workbook(sample_workbook())
    municipalities = (
        Municipality("Divinópolis", "MG", "3122306"),
        Municipality("Itaúna", "MG", "3133808"),
        Municipality("Mateus Leme", "MG", "3140704"),
        Municipality("Juatuba", "MG", "3136652"),
        Municipality("Pará de Minas", "MG", "3147105"),
        Municipality("Ponte Nova", "MG", "3152105"),
    )
    monkeypatch.setattr(
        "services.excel_importer.fetch_state_municipalities",
        lambda state: municipalities,
    )

    snapshot = build_import_snapshot(analysis)

    cities = [item["municipality_name"] for item in snapshot["R.401"]["cities"]]
    assert cities == ["Ponte Nova"]
