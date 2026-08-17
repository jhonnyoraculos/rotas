from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import BinaryIO

import openpyxl
import requests

from utils.city_normalizer import (
    Municipality,
    fetch_state_municipalities,
    identify_municipality,
    normalize_text,
)
from utils.dates import monday_of, today_in_brazil
from utils.route_parser import (
    ParsedRoute,
    extract_route_code,
    is_ignored_city_line,
    strip_route_code,
)

SCHEDULE_SHEET = "CARREGAMENTOS ATUALIZADOS"
ROUTE_CITIES_SHEET = "CIDADES X ROTAS ATUALIZADAS"
WEEKDAY_TOKENS = ("SEGUNDA", "TERCA", "QUARTA", "QUINTA", "SEXTA")


@dataclass
class WorkbookAnalysis:
    sheet_names: list[str]
    schedule_sheet: str
    route_cities_sheet: str
    header_row: int
    schedule: dict[int, list[str]]
    routes: dict[str, ParsedRoute]
    warnings: list[str] = field(default_factory=list)

    def summary_lines(self) -> list[str]:
        lines = [
            f"Abas encontradas: {', '.join(self.sheet_names)}",
            f"Escala: '{self.schedule_sheet}', cabeçalho na linha {self.header_row}",
            f"Relação rota → cidades: '{self.route_cities_sheet}'",
            f"Rotas reconhecidas: {len(self.routes)}",
            "Itens por dia: "
            + ", ".join(
                f"{WEEKDAY_TOKENS[index]}={len(self.schedule.get(index, []))}"
                for index in range(5)
            ),
        ]
        for code, route in list(self.routes.items())[:8]:
            preview = ", ".join(route.cities[:5]) or "sem cidades identificadas"
            lines.append(f"{code} — {route.name}: {preview}")
        lines.extend(f"AVISO: {warning}" for warning in self.warnings)
        return lines


def _find_sheet(workbook: openpyxl.Workbook, expected: str) -> str:
    target = normalize_text(expected)
    exact = {normalize_text(name): name for name in workbook.sheetnames}
    if target in exact:
        return exact[target]
    for normalized, original in exact.items():
        if target in normalized or normalized in target:
            return original
    raise ValueError(
        f"A aba '{expected}' não foi encontrada. Abas: {', '.join(workbook.sheetnames)}"
    )


def _cell_lines(value: object) -> list[str]:
    if value is None or isinstance(value, (date, int, float)):
        return []
    return [
        " ".join(line.split()).strip()
        for line in str(value).splitlines()
        if line.strip()
    ]


def _weekday_index(value: object) -> int | None:
    normalized = normalize_text(value)
    for index, token in enumerate(WEEKDAY_TOKENS):
        if token in normalized:
            return index
    return None


def parse_schedule_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[int, dict[int, list[str]], dict[str, str]]:
    best_row = 0
    best_columns: dict[int, int] = {}
    scan_limit = min(worksheet.max_row, 80)
    for row_number in range(1, scan_limit + 1):
        columns: dict[int, int] = {}
        for cell in worksheet[row_number]:
            weekday = _weekday_index(cell.value)
            if weekday is not None and weekday not in columns:
                columns[weekday] = cell.column
        if len(columns) > len(best_columns):
            best_row, best_columns = row_number, columns
        if len(best_columns) == 5:
            break
    if len(best_columns) < 3:
        raise ValueError(
            "Não foi possível localizar o cabeçalho com os dias úteis na escala."
        )

    schedule: dict[int, list[str]] = {index: [] for index in range(5)}
    names: dict[str, str] = {}
    for weekday, column in best_columns.items():
        for row_number in range(best_row + 1, worksheet.max_row + 1):
            for line in _cell_lines(worksheet.cell(row_number, column).value):
                code = extract_route_code(line)
                if not code:
                    continue
                if code not in schedule[weekday]:
                    schedule[weekday].append(code)
                name = strip_route_code(line)
                if name:
                    names[code] = name
    return best_row, schedule, names


def parse_route_cities_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> dict[str, ParsedRoute]:
    routes: dict[str, ParsedRoute] = {}
    for column in range(1, worksheet.max_column + 1):
        current: ParsedRoute | None = None
        for row_number in range(1, worksheet.max_row + 1):
            lines = _cell_lines(worksheet.cell(row_number, column).value)
            for line in lines:
                code = extract_route_code(line)
                if code:
                    name = strip_route_code(line) or code
                    current = routes.setdefault(
                        code, ParsedRoute(code=code, name=name, original=line)
                    )
                    if current.name == current.code and name != code:
                        current.name = name
                    continue
                if current is None or is_ignored_city_line(line):
                    continue
                normalized = normalize_text(line)
                if normalized and all(
                    normalize_text(existing) != normalized
                    for existing in current.cities
                ):
                    current.cities.append(line)
    return routes


def analyze_workbook(source: str | Path | BinaryIO) -> WorkbookAnalysis:
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=False)
    schedule_sheet = _find_sheet(workbook, SCHEDULE_SHEET)
    cities_sheet = _find_sheet(workbook, ROUTE_CITIES_SHEET)
    header_row, schedule, schedule_names = parse_schedule_sheet(
        workbook[schedule_sheet]
    )
    routes = parse_route_cities_sheet(workbook[cities_sheet])
    warnings: list[str] = []
    for code, name in schedule_names.items():
        if code not in routes:
            routes[code] = ParsedRoute(
                code=code, name=name, original=f"{name} ({code})"
            )
            warnings.append(f"{code} aparece na escala, mas não na relação de cidades.")
        elif routes[code].name == routes[code].code and name:
            routes[code].name = name
    scheduled_codes = {code for values in schedule.values() for code in values}
    for code in sorted(scheduled_codes):
        if not routes[code].cities:
            warnings.append(
                f"{code} não possui linhas de cidade; o nome da rota será validado no IBGE."
            )
    return WorkbookAnalysis(
        sheet_names=list(workbook.sheetnames),
        schedule_sheet=schedule_sheet,
        route_cities_sheet=cities_sheet,
        header_row=header_row,
        schedule=schedule,
        routes=routes,
        warnings=warnings,
    )


def build_import_snapshot(
    analysis: WorkbookAnalysis, state: str = "MG"
) -> dict[str, dict]:
    municipalities: tuple[Municipality, ...] | None
    try:
        municipalities = fetch_state_municipalities(state)
    except (requests.RequestException, ValueError, KeyError):
        municipalities = None
        analysis.warnings.append(
            "IBGE indisponível: cidades foram importadas como pendentes de revisão."
        )

    result: dict[str, dict] = {}
    for code, parsed in analysis.routes.items():
        candidate_cities = list(parsed.cities)
        route_municipality = identify_municipality(parsed.name, state, municipalities)
        if route_municipality and all(
            normalize_text(city) != normalize_text(route_municipality.name)
            for city in candidate_cities
        ):
            candidate_cities.insert(0, route_municipality.name)
            parsed.cities.insert(0, route_municipality.name)
            analysis.warnings = [
                warning
                for warning in analysis.warnings
                if not warning.startswith(f"{code} não possui linhas de cidade")
            ]
        city_rows = []
        for city in candidate_cities:
            municipality = identify_municipality(city, state, municipalities)
            city_rows.append(
                {
                    "city_original": city,
                    "municipality_name": municipality.name if municipality else None,
                    "state": state,
                    "ibge_code": municipality.ibge_code if municipality else None,
                    "needs_review": municipality is None,
                }
            )
        result[code] = {"name": parsed.name, "cities": city_rows}
    return result


def import_workbook(
    source: str | Path | BinaryIO, reference_date: date | None = None
) -> WorkbookAnalysis:
    from services.database import import_snapshot

    analysis = analyze_workbook(source)
    routes = build_import_snapshot(analysis)
    monday = monday_of(reference_date or today_in_brazil())
    import_snapshot(routes, analysis.schedule, monday)
    return analysis


def auto_import_if_available(
    path: str | Path | None = None,
) -> WorkbookAnalysis | None:
    from services.database import count_routes

    if count_routes():
        return None
    candidates = (
        [Path(path)]
        if path is not None
        else [Path("data/ROTAS_2026.xlsx"), Path("ROTAS_2026.xlsx")]
    )
    for candidate in candidates:
        if candidate.exists():
            return import_workbook(candidate)
    return None
