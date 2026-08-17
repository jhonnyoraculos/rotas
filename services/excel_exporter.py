from __future__ import annotations

from collections.abc import Iterable
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.dates import WEEKDAY_NAMES, business_week, week_title


def export_week_to_excel(
    monday: date, schedule: dict[date, list], matches: Iterable
) -> bytes:
    days = business_week(monday)
    match_map: dict[tuple[date, int], list] = {}
    for match in matches:
        match_map.setdefault((match.date, match.route_id), []).append(match)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ESCALA SEMANAL"
    sheet.merge_cells("A1:E1")
    sheet["A1"] = week_title(monday)
    sheet["A1"].font = Font(name="Calibri", size=12, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    alert_fill = PatternFill("solid", fgColor="F4CCCC")
    alert_font = Font(name="Calibri", size=11, color="9C0006", bold=True)

    for column, day in enumerate(days, start=1):
        header = sheet.cell(2, column)
        header.value = f"{WEEKDAY_NAMES[column - 1]}\n{day:%d/%m/%Y}"
        header.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        header.font = Font(name="Calibri", size=11, bold=True)
        header.border = border
        header.fill = (
            alert_fill if any(key[0] == day for key in match_map) else header_fill
        )
        sheet.column_dimensions[get_column_letter(column)].width = 31
    sheet.row_dimensions[2].height = 34

    max_rows = max((len(schedule.get(day, [])) for day in days), default=0)
    for row_index in range(max_rows):
        excel_row = row_index + 3
        sheet.row_dimensions[excel_row].height = 24
        for column, day in enumerate(days, start=1):
            cell = sheet.cell(excel_row, column)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            routes = schedule.get(day, [])
            if row_index >= len(routes):
                continue
            route = routes[row_index]
            affected = match_map.get((day, route.id), [])
            cell.value = f"⚠ {route.label}" if affected else route.label
            if affected:
                cell.fill = alert_fill
                cell.font = alert_font
                details = "\n\n".join(
                    f"{item.city} — {item.name} ({item.holiday_type})"
                    for item in affected
                )
                cell.comment = Comment(details, "Sistema de Rotas")

    sheet.freeze_panes = "A3"
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
